"""
Final parser: correctly handles multi-row category names in the Excel FAQ file.
Produces clean qna_data.json for the WhatsApp chatbot.
"""
import openpyxl
import json
import sys

# Force UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('FAQs with Answers- Complaint.xlsx')
ws = wb.active

# Step 1: Build category start rows with full names
b_rows = []
for row_num in range(2, ws.max_row + 1):
    val_a = ws.cell(row=row_num, column=1).value
    val_b = ws.cell(row=row_num, column=2).value
    if val_b is not None:
        b_rows.append((row_num, val_a, str(val_b).strip()))

cat_start_rows = []
i = 0
while i < len(b_rows):
    row_num, col_a, name_part = b_rows[i]
    if col_a is not None and str(col_a).strip() != '' and str(col_a) != 'No.':
        full_name = name_part
        j = i + 1
        while j < len(b_rows) and b_rows[j][1] is None:
            next_row, _, next_part = b_rows[j]
            if next_row - b_rows[j-1][0] <= 2:
                if full_name.endswith('/') or full_name.endswith('&') or full_name.endswith('& '):
                    full_name = full_name.rstrip() + ' ' + next_part
                else:
                    full_name = full_name + ' ' + next_part
                j += 1
            else:
                break
        full_name = ' '.join(full_name.strip().split())
        cat_start_rows.append((row_num, full_name))
        i = j
    else:
        i += 1

print("=== Categories found ===")
for row_num, name in cat_start_rows:
    print(f"  Row {row_num:4d}: {name}")

# Step 2: Determine category for each row
def get_category_for_row(row_num):
    result = None
    for cat_row, cat_name in cat_start_rows:
        if cat_row <= row_num:
            result = cat_name
        else:
            break
    return result

# Step 3: Clean encoding artifacts
def clean_text(text):
    if not text:
        return text
    text = str(text).strip()
    text = text.replace('\xa0', ' ')
    # Windows-1252 encoding artifacts
    text = text.replace('\x92', "'")
    text = text.replace('\xc6', "'")
    text = text.replace('\xfb', '–')
    text = text.replace('\xf4', '"')
    text = text.replace('\xf6', '"')
    text = text.replace('\x93', '"')
    text = text.replace('\x94', '"')
    text = text.replace('\x96', '–')
    text = text.replace('\x97', '—')
    # Unicode dashes
    text = text.replace('û', '–')
    text = text.replace('ô', '"')
    text = text.replace('ö', '"')
    text = text.replace('Æ', "'")
    return text

# Step 4: Parse Q&A entries
entries = []
current_category = None
current_question = None
current_answer_lines = []

def save_current():
    global current_question, current_answer_lines
    if current_question and current_category and current_answer_lines:
        full_answer = "\n".join(current_answer_lines).strip()
        entries.append({
            "category": current_category,
            "question": current_question,
            "answer": full_answer
        })
    current_question = None
    current_answer_lines = []

for row_num in range(2, ws.max_row + 1):
    col_a = ws.cell(row=row_num, column=1).value  # category serial
    col_c = ws.cell(row=row_num, column=3).value  # question number
    col_d = ws.cell(row=row_num, column=4).value  # question text or answer line
    
    # IMPORTANT: Update category BEFORE processing questions
    # so the first question of a new category gets the right category
    cat = get_category_for_row(row_num)
    if cat:
        current_category = cat
    
    # Check if this is a question row (col C has a number)
    if col_c is not None and col_d is not None:
        try:
            q_num = int(col_c) if not isinstance(col_c, int) else col_c
            if isinstance(q_num, (int, float)):
                save_current()
                current_question = clean_text(col_d)
                continue
        except (ValueError, TypeError):
            pass
    
    # Otherwise, col D is an answer line
    if col_d is not None and str(col_d).strip():
        line = clean_text(col_d)
        current_answer_lines.append(line)

# Save last entry
save_current()

# Remove "Non CPA" category if present (not a real consumer category)
# Actually keep it - it's a valid category in the NGO's system

# Remove duplicate entries
seen = set()
unique_entries = []
for e in entries:
    key = (e["category"], e["question"])
    if key not in seen:
        seen.add(key)
        unique_entries.append(e)
    else:
        print(f"  Duplicate removed: [{e['category']}] {e['question'][:60]}")
entries = unique_entries

# Print summary
categories = {}
for e in entries:
    cat = e["category"]
    if cat not in categories:
        categories[cat] = 0
    categories[cat] += 1

print(f"\nTotal Q&A entries: {len(entries)}")
print(f"Total categories: {len(categories)}")
print("\nCategories and question counts:")
for cat, count in sorted(categories.items()):
    print(f"  {cat}: {count} questions")

# Save to JSON
with open('qna_data.json', 'w', encoding='utf-8') as f:
    json.dump(entries, f, indent=2, ensure_ascii=False)
print(f"\n✅ Saved {len(entries)} entries to qna_data.json")

# Quick verification: show first Q from each category
print("\n--- First Q per category ---")
shown_cats = set()
for e in entries:
    if e["category"] not in shown_cats:
        shown_cats.add(e["category"])
        print(f"\n  [{e['category']}]")
        print(f"    Q: {e['question'][:100]}")
        print(f"    A: {e['answer'][:120]}...")
